#!/usr/bin/env python3
"""
Seed consultant types from Excel file or fallback to hardcoded data.

Usage:
    python backend/scripts/seed_consultant_types.py --dry-run  # Print data without inserting
    python backend/scripts/seed_consultant_types.py            # Insert into database
"""
import argparse
import sys
from pathlib import Path

# Add parent directory to path to import app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models.inspection import ConsultantType

# Hardcoded consultant types as fallback or primary data source
# Based on specification: 21 consultant types with 1-7 inspection stages
CONSULTANT_TYPES = [
    {"name": "Agronomist", "name_he": "אגרונום", "stage_count": 3, "description": "Agricultural and landscaping supervision"},
    {"name": "Soil Engineer", "name_he": "מהנדס קרקע", "stage_count": 4, "description": "Soil testing and foundation supervision"},
    {"name": "Hydrologist", "name_he": "הידרולוג", "stage_count": 2, "description": "Water systems and drainage supervision"},
    {"name": "Waterproofing", "name_he": "איטום", "stage_count": 5, "description": "Waterproofing and sealing supervision"},
    {"name": "Structural Engineer", "name_he": "מהנדס קונסטרוקציה", "stage_count": 7, "description": "Structural integrity supervision"},
    {"name": "Architect", "name_he": "אדריכל", "stage_count": 6, "description": "Architectural design and execution supervision"},
    {"name": "Electrical Engineer", "name_he": "מהנדס חשמל", "stage_count": 5, "description": "Electrical systems supervision"},
    {"name": "Plumbing Engineer", "name_he": "מהנדס אינסטלציה", "stage_count": 4, "description": "Plumbing and water systems supervision"},
    {"name": "HVAC Engineer", "name_he": "מהנדס מיזוג אוויר", "stage_count": 4, "description": "Heating, ventilation, and air conditioning supervision"},
    {"name": "Safety Engineer", "name_he": "מהנדס בטיחות", "stage_count": 5, "description": "Construction safety supervision"},
    {"name": "Accessibility Consultant", "name_he": "יועץ נגישות", "stage_count": 3, "description": "Accessibility compliance supervision"},
    {"name": "Traffic Engineer", "name_he": "מהנדס תנועה", "stage_count": 3, "description": "Traffic and transportation supervision"},
    {"name": "Lighting Designer", "name_he": "מתכנן תאורה", "stage_count": 3, "description": "Lighting systems supervision"},
    {"name": "Signage Consultant", "name_he": "יועץ שילוט", "stage_count": 2, "description": "Signage and wayfinding supervision"},
    {"name": "Radiation Safety", "name_he": "בטיחות קרינה", "stage_count": 2, "description": "Radiation protection supervision"},
    {"name": "Aluminum Specialist", "name_he": "מומחה אלומיניום", "stage_count": 3, "description": "Aluminum work supervision"},
    {"name": "Acoustics Engineer", "name_he": "מהנדס אקוסטיקה", "stage_count": 3, "description": "Acoustic design supervision"},
    {"name": "Green Building Consultant", "name_he": "יועץ בנייה ירוקה", "stage_count": 4, "description": "Sustainable building supervision"},
    {"name": "Development Supervisor", "name_he": "מפקח פיתוח", "stage_count": 5, "description": "Site development supervision"},
    {"name": "Interior Designer", "name_he": "מעצב פנים", "stage_count": 4, "description": "Interior design supervision"},
    {"name": "Elevator Engineer", "name_he": "מהנדס מעליות", "stage_count": 4, "description": "Elevator systems supervision"},
]


def parse_excel_file(file_path: str) -> list[dict]:
    """
    Parse Excel file to extract consultant types.

    Expected Excel structure:
    - Column A: Consultant Type Name (Hebrew/English)
    - Column B: Number of Inspection Stages (1-7)
    - Additional columns may contain stage descriptions

    Args:
        file_path: Path to Excel file

    Returns:
        List of dictionaries with consultant type data
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl==3.1.5")
        sys.exit(1)

    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        consultant_types = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            name = str(row[0]).strip()
            stage_count = int(row[1]) if row[1] else 3  # Default to 3 stages

            # Validate stage count is between 1-7
            if stage_count < 1 or stage_count > 7:
                print(f"Warning: Invalid stage count {stage_count} for {name}, defaulting to 3")
                stage_count = 3

            consultant_types.append({
                "name": name,
                "stage_count": stage_count,
                "description": f"Supervision for {name}"
            })

        wb.close()
        return consultant_types

    except FileNotFoundError:
        print(f"Excel file not found: {file_path}")
        return None
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        return None


def get_consultant_types_data(excel_path: str = None) -> list[dict]:
    """
    Get consultant types data from Excel file or fallback to hardcoded data.

    Args:
        excel_path: Optional path to Excel file

    Returns:
        List of consultant type dictionaries
    """
    # Try to parse Excel file if path provided
    if excel_path:
        excel_data = parse_excel_file(excel_path)
        if excel_data:
            print(f"Loaded {len(excel_data)} consultant types from Excel file")
            return excel_data

    # Try common Excel file locations
    common_paths = [
        "פיקוחים עליונים - כמות בדיקות.xlsx",
        "./פיקוחים עליונים - כמות בדיקות.xlsx",
        "../פיקוחים עליונים - כמות בדיקות.xlsx",
    ]

    for path in common_paths:
        if Path(path).exists():
            excel_data = parse_excel_file(path)
            if excel_data:
                print(f"Loaded {len(excel_data)} consultant types from {path}")
                return excel_data

    # Fallback to hardcoded data
    print("Excel file not found, using hardcoded consultant types")

    # Format hardcoded data
    formatted_data = []
    for item in CONSULTANT_TYPES:
        formatted_data.append({
            "name": item["name_he"],  # Use Hebrew name as primary
            "stage_count": item["stage_count"],
            "description": item["description"]
        })

    return formatted_data


def print_consultant_types(consultant_types: list[dict]):
    """Print consultant types in a readable format."""
    print(f"\n{'='*80}")
    print(f"Consultant Types to Seed: {len(consultant_types)}")
    print(f"{'='*80}\n")

    for idx, ct in enumerate(consultant_types, 1):
        print(f"{idx:2d}. {ct['name']:<40} | Stages: {ct['stage_count']} | {ct.get('description', 'N/A')[:40]}")

    print(f"\n{'='*80}")
    print(f"Total: {len(consultant_types)} consultant types")
    print(f"{'='*80}\n")


def seed_database(consultant_types: list[dict], dry_run: bool = False):
    """
    Seed consultant types into the database.

    Args:
        consultant_types: List of consultant type dictionaries
        dry_run: If True, only print data without inserting
    """
    if dry_run:
        print("DRY RUN MODE - No data will be inserted into database")
        print_consultant_types(consultant_types)
        return

    # Get database settings
    settings = get_settings()

    # Create synchronous engine for seeding script
    engine = create_engine(settings.database_url_sync, echo=False)

    try:
        with Session(engine) as session:
            # Check existing consultant types
            existing = session.execute(select(ConsultantType)).scalars().all()
            existing_names = {ct.name for ct in existing}

            print(f"Found {len(existing)} existing consultant types in database")

            inserted_count = 0
            skipped_count = 0

            for ct_data in consultant_types:
                name = ct_data["name"]

                if name in existing_names:
                    print(f"Skipping '{name}' - already exists")
                    skipped_count += 1
                    continue

                # Create new consultant type
                consultant_type = ConsultantType(
                    name=name,
                    stage_count=ct_data["stage_count"],
                    description=ct_data.get("description"),
                    is_active=True
                )

                session.add(consultant_type)
                inserted_count += 1
                print(f"Inserted: {name} (stages: {ct_data['stage_count']})")

            # Commit changes
            session.commit()

            print(f"\n{'='*80}")
            print("Seeding completed:")
            print(f"  - Inserted: {inserted_count}")
            print(f"  - Skipped:  {skipped_count}")
            print(f"  - Total:    {inserted_count + skipped_count}")
            print(f"{'='*80}\n")

    except Exception as e:
        print(f"Error seeding database: {e}")
        sys.exit(1)
    finally:
        engine.dispose()


def main():
    """Main entry point for the seeding script."""
    parser = argparse.ArgumentParser(
        description="Seed consultant types from Excel file or hardcoded data"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print data without inserting into database"
    )
    parser.add_argument(
        "--excel",
        type=str,
        help="Path to Excel file (optional)"
    )

    args = parser.parse_args()

    # Get consultant types data
    consultant_types = get_consultant_types_data(args.excel)

    if not consultant_types:
        print("Error: No consultant types data available")
        sys.exit(1)

    # Seed database
    seed_database(consultant_types, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
