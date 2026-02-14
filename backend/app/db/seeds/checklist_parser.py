"""Parse checklist Excel files into structured template data.

Handles two formats:
- 'Checklist' sheet (apartment templates): צקליסטים לדירה - לעיון.xlsx
- 'worksheet' sheet (public areas): CL_templates...רמת השרון.xlsx
"""
import glob
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent

APARTMENT_FILE = "צקליסטים לדירה - לעיון.xlsx"
PUBLIC_AREA_GLOB = "CL_templates_*"

ENGLISH_NAMES = {
    "פרוטוקול מסירה לדייר": "Handover Protocol",
    "פרוטוקול פנימי - לפי חללים": "Internal Protocol - By Spaces",
    "תיק דייר": "Resident File",
    "לובי קומתי": "Floor Lobby",
    "פרוטוקול קבלת חזקה בדירה": "Possession Protocol",
    "שטחים ציבוריים": "Public Areas",
}


def parse_flag(val) -> bool:
    if not val:
        return False
    return str(val).strip().lower() in ("v", "true", "1", "yes")


def parse_checklist_sheet(filepath: str, sheet_name: str) -> list[dict]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    templates: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        level = str(row[0]).strip() if row[0] else ""
        if not level or level == "CL Level" or level == "CL level":
            continue

        group = str(row[1]).strip() if row[1] else ""
        cl_name = str(row[2]).strip() if row[2] else ""
        subsection = str(row[4]).strip() if row[4] else ""
        item_name = str(row[5]).strip() if row[5] else ""
        item_category = str(row[6]).strip() if row[6] and str(row[6]).strip() != "None" else None

        if not cl_name or not item_name:
            continue

        must_image_idx = 12 if len(row) > 12 else -1
        must_note_idx = 13 if len(row) > 13 else -1
        must_sig_idx = 14 if len(row) > 14 else -1

        must_image = parse_flag(row[must_image_idx]) if must_image_idx >= 0 else False
        must_note = parse_flag(row[must_note_idx]) if must_note_idx >= 0 else False
        must_signature = parse_flag(row[must_sig_idx]) if must_sig_idx >= 0 else False

        if cl_name not in templates:
            en_name = ENGLISH_NAMES.get(cl_name, cl_name)
            templates[cl_name] = {
                "name": cl_name,
                "level": level,
                "group": group,
                "extra_data": {"name_en": en_name},
                "subsections": {},
            }

        tpl = templates[cl_name]
        if subsection not in tpl["subsections"]:
            tpl["subsections"][subsection] = {
                "name": subsection,
                "order": len(tpl["subsections"]) + 1,
                "items": [],
            }

        tpl["subsections"][subsection]["items"].append({
            "name": item_name,
            "category": item_category,
            "must_image": must_image,
            "must_note": must_note,
            "must_signature": must_signature,
        })

    wb.close()
    return list(templates.values())


def load_all_checklist_data() -> list[dict]:
    all_templates = []

    apartment_path = PROJECT_ROOT / APARTMENT_FILE
    if apartment_path.exists():
        parsed = parse_checklist_sheet(str(apartment_path), "Checklist")
        all_templates.extend(parsed)
        print(f"Parsed {len(parsed)} templates from {APARTMENT_FILE}")
    else:
        print(f"WARNING: {APARTMENT_FILE} not found at {apartment_path}")

    public_files = glob.glob(str(PROJECT_ROOT / PUBLIC_AREA_GLOB))
    if public_files:
        parsed = parse_checklist_sheet(public_files[0], "worksheet")
        all_templates.extend(parsed)
        print(f"Parsed {len(parsed)} templates from {Path(public_files[0]).name}")
    else:
        print(f"WARNING: No {PUBLIC_AREA_GLOB} file found in {PROJECT_ROOT}")

    return all_templates
