"""
Seed script for checklist templates from Excel data.

This script parses the Excel file 'צקליסטים לדירה - לעיון.xlsx' and populates
the database with 5 apartment checklist templates containing 321 total items.
"""
import asyncio
import logging
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Translation dictionary for Hebrew to English
TRANSLATIONS = {
    "פרוטוקול מסירה לדייר": "Handover Protocol to Tenant",
    "פרוטוקול פנימי - לפי חללים": "Internal Protocol - By Spaces",
    "תיק דייר": "Resident File",
    "לובי קומתי": "Floor Lobby",
    "פרוטוקול קבלת חזקה בדירה": "Apartment Possession Protocol",
    "כניסה": "Entrance",
    "מטבח": "Kitchen",
    "סלון ומעברים": "Living Room & Hallways",
    "ממד": "Safe Room",
    "חדר רחצה": "Bathroom",
    "חדרים": "Bedrooms",
    "מרפסות": "Balconies",
    "מסירות": "Handovers",
    "מסירות - פנימי": "Handovers - Internal",
}

# Group name mappings based on template name
GROUP_MAPPINGS = {
    "פרוטוקול מסירה לדייר": "מסירות",
    "פרוטוקול פנימי - לפי חללים": "מסירות - פנימי",
    "תיק דייר": "מסירות",
    "לובי קומתי": "מסירות",
    "פרוטוקול קבלת חזקה בדירה": "מסירות",
}


def translate_to_english(hebrew_text: str) -> str:
    """Translate Hebrew text to English, fallback to original text if no translation."""
    if not hebrew_text:
        return ""
    return TRANSLATIONS.get(hebrew_text.strip(), hebrew_text)


def is_section_header(row: tuple[Any, ...]) -> bool:
    """
    Determine if a row is a section header.
    Section headers typically have text in the first column but different formatting.
    This is a simplified check - adjust based on actual Excel structure.
    """
    # Check if first cell has content and potentially other markers
    # This may need adjustment based on actual Excel structure
    if not row or not row[0]:
        return False

    # Simple heuristic: if only first column has content, it might be a section
    # Or if the row has specific markers. Adjust as needed.
    first_cell = str(row[0]).strip()

    # Skip empty rows
    if not first_cell:
        return False

    # Known section names from spec
    known_sections = [
        "כניסה", "מטבח", "סלון ומעברים", "ממד",
        "חדר רחצה", "חדרים", "מרפסות"
    ]

    return first_cell in known_sections


def parse_excel_templates(excel_path: str) -> list[dict]:
    """
    Parse Excel file and extract template data.

    Expected Excel structure:
    - Each sheet represents a template
    - Sheet name is the template name (Hebrew)
    - Rows contain either section headers or checklist items
    - First row may be headers (skipped)

    Returns:
        List of template dictionaries with nested sections and items
    """
    logger.info(f"Parsing Excel file: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    templates = []

    for sheet_name in wb.sheetnames:
        logger.info(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]

        template_name_he = sheet_name.strip()
        template_data = {
            "name": translate_to_english(template_name_he),
            "name_he": template_name_he,
            "group_name": GROUP_MAPPINGS.get(template_name_he, "מסירות"),
            "level": "project",
            "is_active": True,
            "sections": []
        }

        current_section = None
        item_count = 0

        # Iterate through rows (skip first row if it's a header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if not any(row):
                continue

            first_cell = str(row[0]).strip() if row[0] else ""

            if not first_cell:
                continue

            # Check if this is a section header
            if is_section_header(row):
                # Save previous section if exists
                if current_section:
                    template_data["sections"].append(current_section)

                # Start new section
                current_section = {
                    "name": translate_to_english(first_cell),
                    "name_he": first_cell,
                    "items": []
                }
                logger.debug(f"  New section: {first_cell}")
            else:
                # This is an item row
                # If no section has been created yet, create a default one
                if current_section is None:
                    current_section = {
                        "name": "General",
                        "name_he": "כללי",
                        "items": []
                    }

                # Add item to current section
                current_section["items"].append({
                    "name": translate_to_english(first_cell),
                    "name_he": first_cell,
                })
                item_count += 1

        # Add last section if exists
        if current_section:
            template_data["sections"].append(current_section)

        logger.info(f"  Template '{template_name_he}': {len(template_data['sections'])} sections, {item_count} items")
        templates.append(template_data)

    wb.close()
    return templates


async def create_template_hierarchy(session: AsyncSession, template_data: dict) -> None:
    """
    Create template with nested sections and items in the database.

    Args:
        session: Async database session
        template_data: Dictionary containing template, sections, and items data
    """
    # Import models here to avoid circular imports
    from app.models.checklist import (
        ChecklistTemplate,
        ChecklistSubSection,
        ChecklistItemTemplate
    )

    logger.info(f"Creating template: {template_data['name_he']}")

    # Create template
    template = ChecklistTemplate(
        name=template_data["name"],
        name_he=template_data["name_he"],
        level=template_data["level"],
        group_name=template_data["group_name"],
        is_active=template_data["is_active"]
    )
    session.add(template)
    await session.flush()  # Get template.id

    # Create sub-sections
    for section_idx, section_data in enumerate(template_data["sections"], start=1):
        section = ChecklistSubSection(
            template_id=template.id,
            name=section_data["name"],
            name_he=section_data["name_he"],
            order=section_idx
        )
        session.add(section)
        await session.flush()  # Get section.id

        # Create items
        for item_idx, item_data in enumerate(section_data["items"], start=1):
            item = ChecklistItemTemplate(
                sub_section_id=section.id,
                name=item_data["name"],
                name_he=item_data["name_he"],
                order=item_idx,
                must_image=False,
                must_note=False,
                must_signature=False,
                file_names=[],
                additional_config={}
            )
            session.add(item)

    logger.info(f"  Created {len(template_data['sections'])} sections with items")


async def seed_checklist_templates() -> None:
    """
    Main seed function to populate checklist templates from Excel data.

    This function:
    1. Checks if templates already exist (idempotent)
    2. Locates and parses the Excel file
    3. Creates templates, sections, and items in the database
    """
    try:
        logger.info("Starting checklist template seeding")

        async with AsyncSessionLocal() as session:
            try:
                # Import models here to avoid circular imports
                from app.models.checklist import ChecklistTemplate

                # Check if data already exists (idempotency)
                result = await session.execute(
                    select(ChecklistTemplate).limit(1)
                )
                existing = result.scalar_one_or_none()

                if existing:
                    logger.info("Checklist templates already seeded")
                    return

                # Locate Excel file (should be in project root, 4 directories up from this file)
                excel_path = Path(__file__).parent.parent.parent.parent / "צקליסטים לדירה - לעיון.xlsx"

                if not excel_path.exists():
                    raise FileNotFoundError(
                        f"Excel file not found at: {excel_path}\n"
                        f"Please ensure 'צקליסטים לדירה - לעיון.xlsx' exists in the project root."
                    )

                # Parse Excel file
                templates = parse_excel_templates(str(excel_path))
                logger.info(f"Parsed {len(templates)} templates from Excel")

                # Create templates in database
                for template_data in templates:
                    await create_template_hierarchy(session, template_data)

                await session.commit()
                logger.info(f"Successfully seeded {len(templates)} checklist templates")

            except Exception as e:
                await session.rollback()
                logger.error(f"Error during seeding: {e}")
                raise

    except FileNotFoundError as e:
        logger.error(f"Excel file not found: {e}")
        raise
    except Exception as e:
        logger.error(f"Error seeding checklist templates: {e}")
        raise


def main():
    """Entry point for running seed script directly."""
    asyncio.run(seed_checklist_templates())


if __name__ == "__main__":
    main()
